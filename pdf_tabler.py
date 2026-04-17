import re
import camelot
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font

pdf_path = r"Q4-2025-Supplemental_FINAL.pdf"
output_excel = "Q4-2025-Supplemental_FINAL.xlsx"

def is_blank(value):
    return value is None or str(value).strip() == ""

def split_mixed_percent_amount(value):
    if value is None:
        return None
    s = re.sub(r"\s+", " ", str(value).strip())
    m = re.match(
        r'^(\(?-?\d[\d,]*\.?\d*\)?\s*%)\s*(\$?\s*\(?-?\d[\d,]*(?:\.\d+)?\)?)$',
        s
    )
    if m:
        return [m.group(1).strip(), m.group(2).strip()]
    return None

def repair_row(row):
    row = list(row)

    # turn standalone $ into blank first
    for i in range(len(row)):
        if row[i] is not None and str(row[i]).strip() == "$":
            row[i] = ""

    # split merged percent + amount cells
    for i in range(len(row)):
        split_vals = split_mixed_percent_amount(row[i])
        if not split_vals:
            continue

        prev_blank = i - 1 >= 0 and is_blank(row[i - 1])
        next_blank = i + 1 < len(row) and is_blank(row[i + 1])

        if prev_blank:
            row[i - 1] = split_vals[0]
            row[i] = split_vals[1]
        elif next_blank:
            row[i] = split_vals[0]
            row[i + 1] = split_vals[1]

    return row

def remove_empty_columns(rows, blank_threshold=0.9):
    """
    Remove columns that are:
    1. fully empty, or
    2. mostly empty spacer columns

    Keep all rows, including empty rows.
    """
    if not rows:
        return rows

    max_cols = max(len(row) for row in rows)
    padded_rows = [list(row) + [""] * (max_cols - len(row)) for row in rows]

    keep_col_indexes = []
    row_count = len(padded_rows)

    for col_idx in range(max_cols):
        cleaned_col = []
        for row in padded_rows:
            val = row[col_idx]
            if val is None:
                cleaned_val = ""
            else:
                cleaned_val = str(val).replace("$", "").strip()
            cleaned_col.append(cleaned_val)

        non_blank_count = sum(1 for v in cleaned_col if v != "")
        blank_ratio = 1 - (non_blank_count / row_count)

        # keep only columns that are not mostly blank
        if blank_ratio < blank_threshold:
            keep_col_indexes.append(col_idx)

    cleaned_rows = []
    for row in padded_rows:
        cleaned_rows.append([row[col_idx] for col_idx in keep_col_indexes])

    return cleaned_rows

def clean_cell(value):
    if value is None:
        return value, None

    s = str(value).strip()

    if s == "":
        return "", None

    # remove all dollar signs
    s = s.replace("$", "").strip()

    if s == "":
        return "", None

    s = re.sub(r"\s+%", "%", s)

    if s in {"—", "—%", "— %", "-", "-%", "- %"}:
        return 0, "number"

    if s.endswith("%"):
        s_num = s[:-1].strip()
        if s_num.startswith("(") and s_num.endswith(")"):
            s_num = "-" + s_num[1:-1].strip()
        try:
            return float(s_num.replace(",", "")) / 100, "percent"
        except ValueError:
            return s, None

    s_num = s.replace(",", "").strip()
    if s_num.startswith("(") and s_num.endswith(")"):
        s_num = "-" + s_num[1:-1].strip()

    try:
        return float(s_num), "number"
    except ValueError:
        return s, None

def extract_page_tables(pdf_path, page_num):
    return camelot.read_pdf(
        pdf_path,
        pages=str(page_num),
        flavor="stream",
        row_tol=8,
        edge_tol=300,
        split_text=False,
        strip_text="\n"
    )

reader = PdfReader(pdf_path)
total_pages = len(reader.pages)

wb = Workbook()
wb.remove(wb.active)

for page_num in range(1, total_pages + 1):
    ws = wb.create_sheet(title=f"Page_{page_num}")
    current_row = 1

    try:
        tables = extract_page_tables(pdf_path, page_num)

        if tables.n == 0:
            ws.cell(row=current_row, column=1, value="No table found")
            continue

        for table_index, table in enumerate(tables, start=1):
            ws.cell(row=current_row, column=1, value=f"Table {table_index}")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            # build all rows for this table first
            repaired_rows = []
            for _, row_series in table.df.iterrows():
                row = repair_row(row_series.tolist())
                repaired_rows.append(row)

            # remove empty / mostly empty columns, but keep all rows
            repaired_rows = remove_empty_columns(repaired_rows, blank_threshold=0.9)

            # write cleaned table rows
            for row in repaired_rows:
                for col_idx, raw_value in enumerate(row, start=1):
                    cleaned_value, cell_type = clean_cell(raw_value)
                    cell = ws.cell(row=current_row, column=col_idx, value=cleaned_value)

                    if cell_type == "percent":
                        cell.number_format = "0.0%"
                    elif cell_type == "number":
                        cell.number_format = "#,##0"

                current_row += 1

            # keep spacing between tables
            current_row += 2

        for col_idx in range(1, 20):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else None
            if col_letter:
                ws.column_dimensions[col_letter].width = 36 if col_idx == 1 else 14

    except Exception as e:
        ws.cell(row=current_row, column=1, value=f"Error on page {page_num}: {e}")

wb.save(output_excel)
print(f"Saved to {output_excel}")

# import re
# import camelot
# from pypdf import PdfReader
# from openpyxl import Workbook
# from openpyxl.styles import Font

# pdf_path = r"Q1 2026 Supplemental.pdf"
# output_excel = "Q1 2026 Supplemental.xlsx"

# def is_blank(value):
#     return value is None or str(value).strip() == ""

# def split_mixed_percent_amount(value):
#     if value is None:
#         return None
#     s = re.sub(r"\s+", " ", str(value).strip())
#     m = re.match(
#         r'^(\(?-?\d[\d,]*\.?\d*\)?\s*%)\s*(\$?\s*\(?-?\d[\d,]*(?:\.\d+)?\)?)$',
#         s
#     )
#     if m:
#         return [m.group(1).strip(), m.group(2).strip()]
#     return None

# def repair_row(row):
#     row = list(row)

#     # turn standalone $ into blank first
#     for i in range(len(row)):
#         if row[i] is not None and str(row[i]).strip() == "$":
#             row[i] = ""

#     # split merged percent + amount cells
#     for i in range(len(row)):
#         split_vals = split_mixed_percent_amount(row[i])
#         if not split_vals:
#             continue

#         prev_blank = i - 1 >= 0 and is_blank(row[i - 1])
#         next_blank = i + 1 < len(row) and is_blank(row[i + 1])

#         if prev_blank:
#             row[i - 1] = split_vals[0]
#             row[i] = split_vals[1]
#         elif next_blank:
#             row[i] = split_vals[0]
#             row[i + 1] = split_vals[1]

#     return row

# def remove_empty_columns(rows, blank_threshold=0.9):
#     """
#     Remove columns that are:
#     1. fully empty, or
#     2. mostly empty spacer columns

#     blank_threshold=0.9 means:
#     if 90% or more of the cells in that column are blank, drop it.
#     """
#     if not rows:
#         return rows

#     max_cols = max(len(row) for row in rows)
#     padded_rows = [list(row) + [""] * (max_cols - len(row)) for row in rows]

#     keep_col_indexes = []
#     row_count = len(padded_rows)

#     for col_idx in range(max_cols):
#         cleaned_col = []
#         for row in padded_rows:
#             val = row[col_idx]
#             if val is None:
#                 cleaned_val = ""
#             else:
#                 cleaned_val = str(val).replace("$", "").strip()
#             cleaned_col.append(cleaned_val)

#         non_blank_count = sum(1 for v in cleaned_col if v != "")
#         blank_ratio = 1 - (non_blank_count / row_count)

#         # keep only columns that are not mostly blank
#         if blank_ratio < blank_threshold:
#             keep_col_indexes.append(col_idx)

#     cleaned_rows = []
#     for row in padded_rows:
#         cleaned_rows.append([row[col_idx] for col_idx in keep_col_indexes])

#     return cleaned_rows

# def clean_cell(value):
#     if value is None:
#         return value, None

#     s = str(value).strip()

#     if s == "":
#         return "", None

#     # remove all dollar signs
#     s = s.replace("$", "").strip()

#     if s == "":
#         return "", None

#     s = re.sub(r"\s+%", "%", s)

#     if s in {"—", "—%", "— %", "-", "-%", "- %"}:
#         return 0, "number"

#     if s.endswith("%"):
#         s_num = s[:-1].strip()
#         if s_num.startswith("(") and s_num.endswith(")"):
#             s_num = "-" + s_num[1:-1].strip()
#         try:
#             return float(s_num.replace(",", "")) / 100, "percent"
#         except ValueError:
#             return s, None

#     s_num = s.replace(",", "").strip()
#     if s_num.startswith("(") and s_num.endswith(")"):
#         s_num = "-" + s_num[1:-1].strip()

#     try:
#         return float(s_num), "number"
#     except ValueError:
#         return s, None

# def extract_page_tables(pdf_path, page_num):
#     return camelot.read_pdf(
#         pdf_path,
#         pages=str(page_num),
#         flavor="stream",
#         row_tol=8,
#         edge_tol=300,
#         split_text=False,
#         strip_text="\n"
#     )

# reader = PdfReader(pdf_path)
# total_pages = len(reader.pages)

# wb = Workbook()
# wb.remove(wb.active)

# for page_num in range(1, total_pages + 1):
#     ws = wb.create_sheet(title=f"Page_{page_num}")
#     current_row = 1

#     try:
#         tables = extract_page_tables(pdf_path, page_num)

#         if tables.n == 0:
#             ws.cell(row=current_row, column=1, value="No table found")
#             continue

#         for table_index, table in enumerate(tables, start=1):
#             ws.cell(row=current_row, column=1, value=f"Table {table_index}")
#             ws.cell(row=current_row, column=1).font = Font(bold=True)
#             current_row += 1

#             # build all rows for this table first
#             repaired_rows = []
#             for _, row_series in table.df.iterrows():
#                 row = repair_row(row_series.tolist())
#                 repaired_rows.append(row)

#             # remove fully empty and mostly empty spacer columns
#             repaired_rows = remove_empty_columns(repaired_rows, blank_threshold=0.9)

#             # write cleaned table rows
#             for row in repaired_rows:
#                 for col_idx, raw_value in enumerate(row, start=1):
#                     cleaned_value, cell_type = clean_cell(raw_value)
#                     cell = ws.cell(row=current_row, column=col_idx, value=cleaned_value)

#                     if cell_type == "percent":
#                         cell.number_format = "0.0%"
#                     elif cell_type == "number":
#                         cell.number_format = "#,##0"

#                 current_row += 1

#             current_row += 2

#         for col_idx in range(1, 20):
#             col_letter = chr(64 + col_idx) if col_idx <= 26 else None
#             if col_letter:
#                 ws.column_dimensions[col_letter].width = 36 if col_idx == 1 else 14

#     except Exception as e:
#         ws.cell(row=current_row, column=1, value=f"Error on page {page_num}: {e}")

# wb.save(output_excel)
# print(f"Saved to {output_excel}")

# # import re
# # import camelot
# # from pypdf import PdfReader
# # from openpyxl import Workbook
# # from openpyxl.styles import Font

# # pdf_path = r"Q4-2025-Supplemental_FINAL.pdf"
# # output_excel = "ghgh777777.xlsx"

# # def is_blank(value):
# #     return value is None or str(value).strip() == ""

# # def split_mixed_percent_amount(value):
# #     if value is None:
# #         return None
# #     s = re.sub(r"\s+", " ", str(value).strip())
# #     m = re.match(r'^(\(?-?\d[\d,]*\.?\d*\)?\s*%)\s*(\$?\s*\(?-?\d[\d,]*(?:\.\d+)?\)?)$', s)
# #     if m:
# #         return [m.group(1).strip(), m.group(2).strip()]
# #     return None

# # def repair_row(row):
# #     row = list(row)
# #     for i in range(len(row)):
# #         split_vals = split_mixed_percent_amount(row[i])
# #         if not split_vals:
# #             continue
# #         prev_blank = i - 1 >= 0 and is_blank(row[i - 1])
# #         next_blank = i + 1 < len(row) and is_blank(row[i + 1])
# #         if prev_blank:
# #             row[i - 1] = split_vals[0]
# #             row[i] = split_vals[1]
# #         elif next_blank:
# #             row[i] = split_vals[0]
# #             row[i + 1] = split_vals[1]
# #     return row

# # def clean_cell(value):
# #     if value is None:
# #         return value, None

# #     s = str(value).strip()
# #     if s == "":
# #         return "", None
# #     if s == "$":
# #         return "", None

# #     s = re.sub(r"\s+%", "%", s)

# #     if s in {"—", "—%", "— %", "-", "-%", "- %"}:
# #         return 0, "number"

# #     if s.endswith("%"):
# #         s_num = s[:-1].strip()
# #         if s_num.startswith("(") and s_num.endswith(")"):
# #             s_num = "-" + s_num[1:-1].strip()
# #         try:
# #             return float(s_num.replace(",", "")) / 100, "percent"
# #         except ValueError:
# #             return s, None

# #     if "$" in s:
# #         s_num = s.replace("$", "").replace(",", "").strip()
# #         if s_num.startswith("(") and s_num.endswith(")"):
# #             s_num = "-" + s_num[1:-1].strip()
# #         try:
# #             return float(s_num), "currency"
# #         except ValueError:
# #             return s, None

# #     s_num = s.replace(",", "").strip()
# #     if s_num.startswith("(") and s_num.endswith(")"):
# #         s_num = "-" + s_num[1:-1].strip()
# #     try:
# #         return float(s_num), "number"
# #     except ValueError:
# #         return s, None

# # def extract_page_tables(pdf_path, page_num):
# #     return camelot.read_pdf(
# #         pdf_path,
# #         pages=str(page_num),
# #         flavor="stream",
# #         row_tol=8,
# #         edge_tol=300,
# #         split_text=False,
# #         strip_text="\n"
# #     )

# # reader = PdfReader(pdf_path)
# # total_pages = len(reader.pages)

# # wb = Workbook()
# # wb.remove(wb.active)

# # for page_num in range(1, total_pages + 1):
# #     ws = wb.create_sheet(title=f"Page_{page_num}")
# #     current_row = 1

# #     try:
# #         tables = extract_page_tables(pdf_path, page_num)

# #         if tables.n == 0:
# #             ws.cell(row=current_row, column=1, value="No table found")
# #             continue

# #         for table_index, table in enumerate(tables, start=1):
# #             ws.cell(row=current_row, column=1, value=f"Table {table_index}")
# #             ws.cell(row=current_row, column=1).font = Font(bold=True)
# #             current_row += 1

# #             for _, row_series in table.df.iterrows():
# #                 row = repair_row(row_series.tolist())

# #                 for col_idx, raw_value in enumerate(row, start=1):
# #                     cleaned_value, cell_type = clean_cell(raw_value)
# #                     cell = ws.cell(row=current_row, column=col_idx, value=cleaned_value)

# #                     if cell_type == "percent":
# #                         cell.number_format = "0.0%"
# #                     elif cell_type == "currency":
# #                         cell.number_format = "$#,##0"
# #                     elif cell_type == "number":
# #                         cell.number_format = "#,##0"

# #                 current_row += 1

# #             current_row += 2

# #         for col_idx in range(1, 10):
# #             col_letter = chr(64 + col_idx)
# #             ws.column_dimensions[col_letter].width = 36 if col_idx == 1 else 14

# #     except Exception as e:
# #         ws.cell(row=current_row, column=1, value=f"Error on page {page_num}: {e}")

# # wb.save(output_excel)
# # print(f"Saved to {output_excel}")


# # import re
# # import gc
# # import camelot
# # import pandas as pd
# # from pypdf import PdfReader
# # from openpyxl import load_workbook

# # pdf_path = r"Q4-2025-Supplemental_FINAL.pdf"
# # output_excel = "213232326565ooooooooqqqqqqqqqqq.xlsx"

# # reader = PdfReader(pdf_path)
# # total_pages = len(reader.pages)


# # def clean_cell(value):
# #     if pd.isna(value):
# #         return value, None

# #     s = str(value).strip()

# #     # Treat standalone dollar sign as empty
# #     if s == "$":
# #         return "", None

# #     # Dash / blank numeric-style values -> 0
# #     if s in {"—", "— %", "—%", "-", "- %", "-%", " — %", " —%", "— %", "—"}:
# #         return 0, "number"

# #     # Remove spaces before %
# #     s = re.sub(r"\s+%", "%", s)

# #     # Convert percent text like 3.5% or (0.7)% to real numeric percent
# #     if s.endswith("%"):
# #         s_num = s.replace("%", "").strip()
# #         if s_num.startswith("(") and s_num.endswith(")"):
# #             s_num = "-" + s_num[1:-1].strip()
# #         try:
# #             return float(s_num) / 100, "percent"
# #         except ValueError:
# #             return s, None

# #     # If currency text is already in one cell like "$ 3,231" or "$3,231"
# #     if "$" in s:
# #         s = re.sub(r"\$\s+", "$", s)
# #         s_num = s.replace("$", "").replace(",", "").strip()

# #         if s_num.startswith("(") and s_num.endswith(")"):
# #             s_num = "-" + s_num[1:-1].strip()

# #         try:
# #             return float(s_num), "currency"
# #         except ValueError:
# #             return s, None

# #     # Plain numbers with commas
# #     s_no_comma = s.replace(",", "")
# #     try:
# #         return float(s_no_comma), "number"
# #     except ValueError:
# #         return s, None


# # def blank_out_dollar_cells(df):
# #     """
# #     Turn cells that contain only '$' into blank cells.
# #     """
# #     df = df.copy().astype(object)

# #     for r in range(df.shape[0]):
# #         for c in range(df.shape[1]):
# #             val = "" if pd.isna(df.iat[r, c]) else str(df.iat[r, c]).strip()
# #             if val == "$":
# #                 df.iat[r, c] = ""

# #     return df


# # def drop_fully_empty_columns(df):
# #     """
# #     Delete only columns that are completely empty after cleanup.
# #     """
# #     df = df.copy()

# #     keep_cols = []
# #     for c in range(df.shape[1]):
# #         col_vals = df.iloc[:, c].fillna("").astype(str).str.strip()
# #         if (col_vals != "").any():
# #             keep_cols.append(c)

# #     return df.iloc[:, keep_cols]


# # def bbox_to_table_area(bbox, padding=2):
# #     x1, y1, x2, y2 = bbox
# #     x1 = max(0, x1 - padding)
# #     y1 = max(0, y1 - padding)
# #     x2 = x2 + padding
# #     y2 = y2 + padding
# #     return f"{x1},{y2},{x2},{y1}"


# # def extract_page_tables(pdf_path, page_num):
# #     lattice_tables = camelot.read_pdf(
# #         pdf_path,
# #         pages=str(page_num),
# #         flavor="lattice",
# #         line_scale=40,
# #         joint_tol=3,
# #         line_tol=2,
# #         process_background=False
# #     )

# #     print(f"Page {page_num}: lattice found {lattice_tables.n} table(s)")

# #     if lattice_tables.n == 0:
# #         stream_tables = camelot.read_pdf(
# #             pdf_path,
# #             pages=str(page_num),
# #             flavor="stream",
# #             row_tol=8,
# #             edge_tol=300,
# #             split_text=False,
# #             strip_text="\n"
# #         )
# #     else:
# #         table_areas = [bbox_to_table_area(t._bbox, padding=3) for t in lattice_tables]

# #         stream_tables = camelot.read_pdf(
# #             pdf_path,
# #             pages=str(page_num),
# #             flavor="stream",
# #             table_areas=table_areas,
# #             row_tol=8,
# #             edge_tol=300,
# #             split_text=False,
# #             strip_text="\n"
# #         )

# #     print(f"Page {page_num}: final stream found {stream_tables.n} table(s)")
# #     return stream_tables


# # with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
# #     percent_positions = {}
# #     currency_positions = {}

# #     for page_num in range(1, total_pages + 1):
# #         sheet_name = f"Page_{page_num}"
# #         percent_cells = []
# #         currency_cells = []

# #         try:
# #             stream_tables = extract_page_tables(pdf_path, page_num)

# #             if stream_tables.n > 0:
# #                 start_row = 0

# #                 for table_index, table in enumerate(stream_tables, start=1):
# #                     df = table.df.copy().astype(object)

# #                     # Just ignore standalone $
# #                     df = blank_out_dollar_cells(df)

# #                     # Delete columns that became fully empty
# #                     df = drop_fully_empty_columns(df)

# #                     pd.DataFrame([[f"Table {table_index}"]]).to_excel(
# #                         writer,
# #                         sheet_name=sheet_name,
# #                         index=False,
# #                         header=False,
# #                         startrow=start_row
# #                     )
# #                     start_row += 1

# #                     # Clean and type-convert cells
# #                     for r in range(df.shape[0]):
# #                         for c in range(df.shape[1]):
# #                             cleaned_value, cell_type = clean_cell(df.iat[r, c])
# #                             df.iat[r, c] = cleaned_value

# #                             excel_row = start_row + r + 2
# #                             excel_col = c + 1

# #                             if cell_type == "percent":
# #                                 percent_cells.append((excel_row, excel_col))
# #                             elif cell_type == "currency":
# #                                 currency_cells.append((excel_row, excel_col))

# #                     df.to_excel(
# #                         writer,
# #                         sheet_name=sheet_name,
# #                         index=False,
# #                         startrow=start_row
# #                     )

# #                     start_row += len(df) + 3

# #                 percent_positions[sheet_name] = percent_cells
# #                 currency_positions[sheet_name] = currency_cells

# #             else:
# #                 pd.DataFrame([["No table found"]]).to_excel(
# #                     writer,
# #                     sheet_name=sheet_name,
# #                     index=False,
# #                     header=False
# #                 )

# #         except Exception as e:
# #             pd.DataFrame([[f"Error on page {page_num}: {e}"]]).to_excel(
# #                 writer,
# #                 sheet_name=sheet_name,
# #                 index=False,
# #                 header=False
# #             )

# #         finally:
# #             gc.collect()

# # wb = load_workbook(output_excel)

# # for sheet_name, percent_cells in percent_positions.items():
# #     ws = wb[sheet_name]
# #     for row_num, col_num in percent_cells:
# #         ws.cell(row=row_num, column=col_num).number_format = "0.0%"

# # for sheet_name, currency_cells in currency_positions.items():
# #     ws = wb[sheet_name]
# #     for row_num, col_num in currency_cells:
# #         ws.cell(row=row_num, column=col_num).number_format = "$#,##0"

# # wb.save(output_excel)

# # print("Done")

