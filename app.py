import re
import gc
import os
import tempfile
import streamlit as st
import camelot
import pandas as pd
from pypdf import PdfReader
from openpyxl import load_workbook


def clean_cell(value):
    if pd.isna(value):
        return value, None

    s = str(value).strip()

    if s == "$":
        return "", None

    if s in {"—", "— %", "—%", "-", "- %", "-%", " — %", " —%", "— %"}:
        return 0, "number"

    s = re.sub(r"\s+%", "%", s)

    if s.endswith("%"):
        s_num = s.replace("%", "").strip()
        if s_num.startswith("(") and s_num.endswith(")"):
            s_num = "-" + s_num[1:-1].strip()
        try:
            return float(s_num) / 100, "percent"
        except ValueError:
            return s, None

    if "$" in s:
        s = re.sub(r"\$\s+", "$", s)
        s_num = s.replace("$", "").replace(",", "").strip()

        if s_num.startswith("(") and s_num.endswith(")"):
            s_num = "-" + s_num[1:-1].strip()

        try:
            return float(s_num), "currency"
        except ValueError:
            return s, None

    s_no_comma = s.replace(",", "")
    try:
        return float(s_no_comma), "number"
    except ValueError:
        return s, None


def blank_out_dollar_cells(df):
    df = df.copy().astype(object)

    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            val = "" if pd.isna(df.iat[r, c]) else str(df.iat[r, c]).strip()
            if val == "$":
                df.iat[r, c] = ""

    return df


def drop_fully_empty_columns(df):
    df = df.copy()

    keep_cols = []
    for c in range(df.shape[1]):
        col_vals = df.iloc[:, c].fillna("").astype(str).str.strip()
        if (col_vals != "").any():
            keep_cols.append(c)

    return df.iloc[:, keep_cols]


def bbox_to_table_area(bbox, padding=2):
    x1, y1, x2, y2 = bbox
    x1 = max(0, x1 - padding)
    y1 = max(0, y1 - padding)
    x2 = x2 + padding
    y2 = y2 + padding
    return f"{x1},{y2},{x2},{y1}"


def extract_page_tables(pdf_path, page_num):
    lattice_tables = camelot.read_pdf(
        pdf_path,
        pages=str(page_num),
        flavor="lattice",
        line_scale=40,
        joint_tol=3,
        line_tol=2,
        process_background=False
    )

    if lattice_tables.n == 0:
        stream_tables = camelot.read_pdf(
            pdf_path,
            pages=str(page_num),
            flavor="stream",
            row_tol=8,
            edge_tol=300,
            split_text=False,
            strip_text="\n"
        )
    else:
        table_areas = [bbox_to_table_area(t._bbox, padding=3) for t in lattice_tables]

        stream_tables = camelot.read_pdf(
            pdf_path,
            pages=str(page_num),
            flavor="stream",
            table_areas=table_areas,
            row_tol=8,
            edge_tol=300,
            split_text=False,
            strip_text="\n"
        )

    return stream_tables


def convert_pdf_to_excel(pdf_path, output_excel):
    reader = PdfReader(pdf_path)
    total_pages = len(reader.pages)

    percent_positions = {}
    currency_positions = {}

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        for page_num in range(1, total_pages + 1):
            sheet_name = f"Page_{page_num}"
            percent_cells = []
            currency_cells = []

            try:
                stream_tables = extract_page_tables(pdf_path, page_num)

                if stream_tables.n > 0:
                    start_row = 0

                    for table_index, table in enumerate(stream_tables, start=1):
                        df = table.df.copy().astype(object)

                        df = blank_out_dollar_cells(df)
                        df = drop_fully_empty_columns(df)

                        pd.DataFrame([[f"Table {table_index}"]]).to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            header=False,
                            startrow=start_row
                        )
                        start_row += 1

                        for r in range(df.shape[0]):
                            for c in range(df.shape[1]):
                                cleaned_value, cell_type = clean_cell(df.iat[r, c])
                                df.iat[r, c] = cleaned_value

                                excel_row = start_row + r + 2
                                excel_col = c + 1

                                if cell_type == "percent":
                                    percent_cells.append((excel_row, excel_col))
                                elif cell_type == "currency":
                                    currency_cells.append((excel_row, excel_col))

                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            startrow=start_row
                        )

                        start_row += len(df) + 3

                    percent_positions[sheet_name] = percent_cells
                    currency_positions[sheet_name] = currency_cells

                else:
                    pd.DataFrame([["No table found"]]).to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False,
                        header=False
                    )

            except Exception as e:
                pd.DataFrame([[f"Error on page {page_num}: {e}"]]).to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=False
                )

            finally:
                gc.collect()

    wb = load_workbook(output_excel)

    for sheet_name, percent_cells in percent_positions.items():
        ws = wb[sheet_name]
        for row_num, col_num in percent_cells:
            ws.cell(row=row_num, column=col_num).number_format = "0.0%"

    for sheet_name, currency_cells in currency_positions.items():
        ws = wb[sheet_name]
        for row_num, col_num in currency_cells:
            ws.cell(row=row_num, column=col_num).number_format = "$#,##0"

    wb.save(output_excel)


st.title("PDF Table to Excel Converter")

uploaded_file = st.file_uploader("Choose a PDF file", type=["pdf"])

if uploaded_file is not None:
    st.success(f"Selected file: {uploaded_file.name}")

    output_name = os.path.splitext(uploaded_file.name)[0] + ".xlsx"

    if st.button("Convert to Excel"):
        with st.spinner("Converting PDF to Excel..."):
            with tempfile.TemporaryDirectory() as tmpdir:
                pdf_path = os.path.join(tmpdir, uploaded_file.name)
                output_excel = os.path.join(tmpdir, output_name)

                with open(pdf_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                convert_pdf_to_excel(pdf_path, output_excel)

                with open(output_excel, "rb") as f:
                    excel_bytes = f.read()

        st.success("Conversion complete!")

        st.download_button(
            label="Download Excel File",
            data=excel_bytes,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )